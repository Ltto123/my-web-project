"""
不背单词 API 路由 — 单词集管理 + 文件上传解析 + 学习进度
"""
import threading
import re as _re
from datetime import datetime, timezone
import io
import os
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, UploadFile, File, Form, Depends
from sqlalchemy.orm import Session
from sqlalchemy import text

from backend.auth import get_current_user
from backend.vocab_ai import parse_and_complete
import backend.models as models
import backend.schemas as schemas

BLOG_OWNER = os.getenv("BLOG_OWNER_USERNAME", "").strip()


def _is_owner(user) -> bool:
    if not BLOG_OWNER or not user:
        return False
    return user.username == BLOG_OWNER

router = APIRouter(prefix="/api/v1/vocab", tags=["vocab"])

MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB


def _extract_text_pdf(contents: bytes) -> str:
    """PDF 文本提取"""
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(contents))
    pages = []
    for page in reader.pages:
        t = page.extract_text()
        if t and t.strip():
            pages.append(t)
    if not pages:
        raise ValueError("PDF 中未提取到文字内容，请确认 PDF 包含文字而非扫描图片")
    return "\n\n".join(pages)


def _extract_text_docx(contents: bytes) -> str:
    """Word (.docx) 文本提取"""
    from docx import Document
    doc = Document(io.BytesIO(contents))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    if not paragraphs:
        raise ValueError("Word 文档中未提取到文字内容")
    return "\n\n".join(paragraphs)


def _extract_text_xlsx(contents: bytes) -> str:
    """Excel (.xlsx/.xls) 文本提取 — 逐行逐列拼接"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                rows.append("\t".join(cells))
    wb.close()
    if not rows:
        raise ValueError("Excel 文件中未提取到文字内容")
    return "\n".join(rows)


def _extract_text_html(contents: bytes) -> str:
    """HTML 文本提取 — 去标签、去脚本、去样式"""
    text = contents.decode("utf-8", errors="replace")
    # Remove script/style blocks
    text = _re.sub(r'<(script|style)[^>]*>.*?</\1>', ' ', text, flags=_re.DOTALL | _re.IGNORECASE)
    # Remove HTML tags
    text = _re.sub(r'<[^>]+>', ' ', text)
    # Collapse whitespace
    text = _re.sub(r'\s+', ' ', text).strip()
    if not text:
        raise ValueError("HTML 文件中未提取到文字内容")
    return text


def _extract_text_plain(contents: bytes) -> str:
    """纯文本文件 — 自动检测编码"""
    for encoding in ["utf-8", "gbk", "latin-1"]:
        try:
            return contents.decode(encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
    # Last resort
    return contents.decode("utf-8", errors="replace")


# ── 格式分发映射 ──
_EXTRACTORS = {
    ".pdf": _extract_text_pdf,
    ".docx": _extract_text_docx,
    ".xlsx": _extract_text_xlsx,
    ".html": _extract_text_html,
    ".htm": _extract_text_html,
}


def _extract_text(contents: bytes, filename: str) -> str:
    """从任意格式文件提取文本，自动选择合适的解析器"""
    ext = Path(filename).suffix.lower()

    extractor = _EXTRACTORS.get(ext)
    if extractor is not None:
        return extractor(contents)

    # 未识别的格式 → 尝试当成纯文本
    return _extract_text_plain(contents)


def _serialize_word(w: models.VocabWordModel) -> dict:
    return {
        "id": w.id,
        "word": w.word,
        "pos": w.pos,
        "def_en": w.def_en,
        "def_zh": w.def_zh,
        "example_en": w.example_en,
        "example_zh": w.example_zh,
        "is_phrase": w.is_phrase,
        "sort_order": w.sort_order,
    }


def _serialize_set(s: models.VocabSetModel, progress_pct: float = 0) -> dict:
    return {
        "id": s.id,
        "name": s.name,
        "description": s.description,
        "word_count": s.word_count,
        "user_id": s.user_id,
        "created_at": s.created_at.isoformat() if s.created_at else None,
        "progress_pct": progress_pct,
        "status": getattr(s, "status", "completed") or "completed",
    }


# ═══════════════ 单词集管理 ═══════════════


@router.get("/sets", response_model=schemas.HttpResponseSchema)
def list_vocab_sets(
    current_user: Optional[models.UserModel] = Depends(get_current_user),
    db: Session = Depends(models.get_db),
):
    """获取当前用户的单词集列表（只能看到自己上传的）"""
    if not current_user:
        return schemas.HttpResponseSchema(code=0, msg="success", data=[])
    sets = db.query(models.VocabSetModel).filter(
        models.VocabSetModel.user_id == current_user.id
    ).order_by(models.VocabSetModel.id.desc()).all()
    result = []
    for s in sets:
        pct = 0
        if s.word_count > 0:
            learned = db.query(models.VocabProgressModel).filter(
                models.VocabProgressModel.user_id == current_user.id,
                models.VocabProgressModel.word_id.in_(
                    db.query(models.VocabWordModel.id).filter(models.VocabWordModel.set_id == s.id)
                ),
                models.VocabProgressModel.stage >= 1,
            ).count()
            pct = round(learned / s.word_count * 100, 1)
        result.append(_serialize_set(s, pct))
    return schemas.HttpResponseSchema(code=0, msg="success", data=result)


@router.get("/sets/{set_id}", response_model=schemas.HttpResponseSchema)
def get_vocab_set(
    set_id: int,
    current_user: Optional[models.UserModel] = Depends(get_current_user),
    db: Session = Depends(models.get_db),
):
    """获取单词集详情（仅上传者本人可查看）"""
    if not current_user:
        return schemas.HttpResponseSchema(code=403, msg="请先登录", data=None)

    vset = db.query(models.VocabSetModel).filter(models.VocabSetModel.id == set_id).first()
    if not vset:
        return schemas.HttpResponseSchema(code=404, msg="单词集不存在", data=None)
    if vset.user_id != current_user.id:
        return schemas.HttpResponseSchema(code=403, msg="无权查看此单词集", data=None)

    words = db.query(models.VocabWordModel).filter(
        models.VocabWordModel.set_id == set_id
    ).order_by(models.VocabWordModel.sort_order).all()

    return schemas.HttpResponseSchema(code=0, msg="success", data={
        **_serialize_set(vset),
        "words": [_serialize_word(w) for w in words],
    })


def _process_vocab_async(set_id: int, text_content: str, user_id: int):
    """后台线程：调用 DeepSeek AI 解析单词并写入数据库"""
    db = models.SessionLocal()
    try:
        words_data = parse_and_complete(text_content)

        vset = db.query(models.VocabSetModel).filter(models.VocabSetModel.id == set_id).first()
        if not vset:
            return
        # Guard: if set was deleted or status changed (e.g. re-upload), abort
        if vset.status != "processing":
            return

        if not words_data:
            vset.status = "error"
            vset.error_message = "未能从文件中解析出任何单词，请检查文件内容"
            vset.description = "解析失败：无单词数据"
            db.commit()
            return

        # Create word entries
        for wd in words_data:
            word_entry = models.VocabWordModel(
                set_id=vset.id,
                word=wd["word"],
                pos=wd.get("pos"),
                def_en=wd.get("def_en"),
                def_zh=wd.get("def_zh"),
                example_en=wd.get("example_en"),
                example_zh=wd.get("example_zh"),
                is_phrase=wd.get("is_phrase", 0),
                sort_order=wd.get("sort_order", 0),
            )
            db.add(word_entry)

        vset.word_count = len(words_data)
        vset.status = "completed"
        vset.description = f"由用户上传，共 {len(words_data)} 词"
        db.commit()

    except Exception as e:
        try:
            db.rollback()  # 清理失败的事务，否则 session 无法继续查询
            vset = db.query(models.VocabSetModel).filter(models.VocabSetModel.id == set_id).first()
            if vset:
                vset.status = "error"
                vset.error_message = str(e)[:500]
                vset.description = "解析出错"
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


@router.post("/sets/upload", response_model=schemas.HttpResponseSchema)
async def upload_vocab_set(
    file: UploadFile = File(...),
    name: Optional[str] = Form(None),
    current_user: Optional[models.UserModel] = Depends(get_current_user),
    db: Session = Depends(models.get_db),
):
    """上传单词文件，立即返回，AI 在后台线程异步解析"""
    if not current_user:
        return schemas.HttpResponseSchema(code=403, msg="请先登录", data=None)

    if not file.filename:
        return schemas.HttpResponseSchema(code=400, msg="请选择文件", data=None)

    contents = await file.read()
    if not contents:
        return schemas.HttpResponseSchema(code=400, msg="文件为空", data=None)

    if len(contents) > MAX_FILE_SIZE:
        return schemas.HttpResponseSchema(code=400, msg="文件大小不能超过 5MB", data=None)

    # Extract text synchronously (fast — just reads PDF, no AI involved)
    try:
        file_text = _extract_text(contents, file.filename)
    except ValueError as e:
        return schemas.HttpResponseSchema(code=400, msg=str(e), data=None)

    # Create vocab set immediately with "processing" status
    set_name = name or Path(file.filename).stem
    vset = models.VocabSetModel(
        name=set_name,
        description="AI 正在解析中...",
        word_count=0,
        user_id=current_user.id,
        created_at=datetime.now(timezone.utc),
        status="processing",
    )
    db.add(vset)
    db.commit()
    db.refresh(vset)

    # Start background AI processing (daemon thread, auto-killed on server exit)
    thread = threading.Thread(
        target=_process_vocab_async,
        args=(vset.id, file_text, current_user.id),
        daemon=True,
    )
    thread.start()

    return schemas.HttpResponseSchema(code=0, msg="上传成功，AI 正在后台解析", data={
        "set_id": vset.id,
        "name": set_name,
        "status": "processing",
    })


@router.get("/sets/{set_id}/status", response_model=schemas.HttpResponseSchema)
def get_set_status(
    set_id: int,
    current_user: Optional[models.UserModel] = Depends(get_current_user),
    db: Session = Depends(models.get_db),
):
    """轮询单词集的 AI 解析状态（仅上传者本人可查询）"""
    if not current_user:
        return schemas.HttpResponseSchema(code=403, msg="请先登录", data=None)

    vset = db.query(models.VocabSetModel).filter(models.VocabSetModel.id == set_id).first()
    if not vset:
        return schemas.HttpResponseSchema(code=404, msg="单词集不存在", data=None)
    if vset.user_id != current_user.id:
        return schemas.HttpResponseSchema(code=403, msg="无权查看此单词集", data=None)

    return schemas.HttpResponseSchema(code=0, msg="success", data={
        "status": getattr(vset, "status", "completed") or "completed",
        "word_count": vset.word_count,
        "error_message": getattr(vset, "error_message", None),
    })


@router.delete("/sets/{set_id}", response_model=schemas.HttpResponseSchema)
def delete_vocab_set(
    set_id: int,
    current_user: Optional[models.UserModel] = Depends(get_current_user),
    db: Session = Depends(models.get_db),
):
    """删除单词集（仅上传者或博主）"""
    if not current_user:
        return schemas.HttpResponseSchema(code=403, msg="请先登录", data=None)

    vset = db.query(models.VocabSetModel).filter(models.VocabSetModel.id == set_id).first()
    if not vset:
        return schemas.HttpResponseSchema(code=404, msg="单词集不存在", data=None)

    if not _is_owner(current_user) and vset.user_id != current_user.id:
        return schemas.HttpResponseSchema(code=403, msg="无权删除此单词集", data=None)

    db.delete(vset)
    db.commit()
    return schemas.HttpResponseSchema(code=0, msg="删除成功", data=None)


# ═══════════════ 学习进度 ═══════════════


@router.get("/progress/{set_id}", response_model=schemas.HttpResponseSchema)
def get_progress(
    set_id: int,
    current_user: Optional[models.UserModel] = Depends(get_current_user),
    db: Session = Depends(models.get_db),
):
    """获取用户在指定单词集的学习进度"""
    if not current_user:
        return schemas.HttpResponseSchema(code=403, msg="请先登录", data=None)

    word_ids = db.query(models.VocabWordModel.id).filter(
        models.VocabWordModel.set_id == set_id
    ).subquery()

    rows = db.query(models.VocabProgressModel).filter(
        models.VocabProgressModel.user_id == current_user.id,
        models.VocabProgressModel.word_id.in_(word_ids),
    ).all()

    progress_map = {}
    for r in rows:
        progress_map[r.word_id] = {
            "word_id": r.word_id,
            "stage": r.stage,
            "correct_count": r.correct_count,
            "wrong_count": r.wrong_count,
            "spelling_passed": bool(r.spelling_passed),
            "last_reviewed": r.last_reviewed,
        }

    return schemas.HttpResponseSchema(code=0, msg="success", data=progress_map)


@router.post("/progress", response_model=schemas.HttpResponseSchema)
def update_progress(
    body: schemas.VocabProgressUpdate,
    current_user: Optional[models.UserModel] = Depends(get_current_user),
    db: Session = Depends(models.get_db),
):
    """更新单个单词的学习进度"""
    if not current_user:
        return schemas.HttpResponseSchema(code=403, msg="请先登录", data=None)

    import time
    now = int(time.time() * 1000)

    existing = db.query(models.VocabProgressModel).filter(
        models.VocabProgressModel.user_id == current_user.id,
        models.VocabProgressModel.word_id == body.word_id,
    ).first()

    if existing:
        existing.stage = body.stage
        existing.correct_count = body.correct_count
        existing.wrong_count = body.wrong_count
        existing.last_reviewed = now
    else:
        db.add(models.VocabProgressModel(
            user_id=current_user.id,
            word_id=body.word_id,
            stage=body.stage,
            correct_count=body.correct_count,
            wrong_count=body.wrong_count,
            last_reviewed=now,
        ))

    db.commit()
    return schemas.HttpResponseSchema(code=0, msg="success", data=None)


@router.post("/progress/spell", response_model=schemas.HttpResponseSchema)
def mark_spelling(
    body: schemas.VocabSpellUpdate,
    current_user: Optional[models.UserModel] = Depends(get_current_user),
    db: Session = Depends(models.get_db),
):
    """标记单词拼写通过"""
    if not current_user:
        return schemas.HttpResponseSchema(code=403, msg="请先登录", data=None)

    existing = db.query(models.VocabProgressModel).filter(
        models.VocabProgressModel.user_id == current_user.id,
        models.VocabProgressModel.word_id == body.word_id,
    ).first()

    if existing:
        existing.spelling_passed = 1
    else:
        db.add(models.VocabProgressModel(
            user_id=current_user.id,
            word_id=body.word_id,
            spelling_passed=1,
        ))

    db.commit()
    return schemas.HttpResponseSchema(code=0, msg="success", data=None)


@router.post("/progress/reset/{set_id}", response_model=schemas.HttpResponseSchema)
def reset_progress(
    set_id: int,
    current_user: Optional[models.UserModel] = Depends(get_current_user),
    db: Session = Depends(models.get_db),
):
    """重置用户在指定单词集的所有学习进度"""
    if not current_user:
        return schemas.HttpResponseSchema(code=403, msg="请先登录", data=None)

    word_ids = db.query(models.VocabWordModel.id).filter(
        models.VocabWordModel.set_id == set_id
    ).subquery()

    db.query(models.VocabProgressModel).filter(
        models.VocabProgressModel.user_id == current_user.id,
        models.VocabProgressModel.word_id.in_(word_ids),
    ).delete(synchronize_session=False)

    db.commit()
    return schemas.HttpResponseSchema(code=0, msg="进度已重置", data=None)
